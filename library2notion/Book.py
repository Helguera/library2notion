import openpyxl
import epub_meta
from PyPDF2 import PdfReader


class Book:
    def __init__(self, fileName, formats, temp_format):
        self.fileName = fileName
        self.title = ''
        self.priority = ''
        tags = (fileName + temp_format).split('/')[:-1]
        self.tags = tags
        self.status = ''
        self.author = ''
        self.publisher = ''
        self.comments = ''
        self.formats = formats
        self.temp_format = temp_format
        self.isbn = ''
        self.action = 'upload'

    def copyToExcel(self, rowNum, ws):
        ws.cell(row=rowNum, column=1).value = self.fileName
        ws.cell(row=rowNum, column=2).value = self.title    
        ws.cell(row=rowNum, column=3).value = self.priority
        ws.cell(row=rowNum, column=4).value = self.tags
        ws.cell(row=rowNum, column=5).value = self.status
        ws.cell(row=rowNum, column=6).value = self.author
        ws.cell(row=rowNum, column=7).value = self.publisher
        ws.cell(row=rowNum, column=8).value = self.comments
        ws.cell(row=rowNum, column=9).value = ', '.join(self.formats)
        ws.cell(row=rowNum, column=10).value = self.isbn

    def extractMetadata(self):
        for format in self.formats:
            filePath = self.fileName + '.' + format.lower()

            if 'EPUB' in format.upper():
                data = epub_meta.get_epub_metadata(filePath, read_cover_image=True, read_toc=True)

                # TITLE
                self.title = data.title

                # AUTHOR
                self.author = ' '.join(data.authors)

                # PUBLISHER
                if data.publisher:
                    self.publisher = data.publisher.replace(',','')

                # ISBN
                self.isbn = ' '.join(data.identifiers)

            elif 'PDF' in format.upper():
                try:
                    reader = PdfReader(filePath)
                    pdf_info = reader.metadata

                    # TITLE
                    if pdf_info.title: 
                        self.title = pdf_info.title
                    else:
                         self.title = filePath.split('/')[len(filePath.split('/')) - 1].split('.')[0]

                    # AUTHOR
                    author = ''
                    if pdf_info.author is not None:
                        self.author = pdf_info.author

                except Exception as e:
                    if str(e) != 'EOF marker not found':
                        self.status = 'BROKEN!'

        if self.tags[len(self.tags) - 1] == self.title:
            self.tags = self.tags[:-1]
        self.tags = ', '.join(self.tags[1:])