
import openpyxl
import os
import pandas as pd
import json
import collections
import argparse

from .Book import Book
from .ProgressBar import printProgressBar
from .BookCollection import BookCollection


# LINE ARGUMENTS ------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser()

    parser.add_argument("-p", "--Path", help="Origin path", required=True)
    parser.add_argument("-l", "--LogFilePath", help="Log File path", required=False)
    parser.add_argument("-e", "--Extensions", nargs='+', help="Supported Extensions", required=False)

    parser.add_argument("-t", "--NotionToken", help="Notion Token", required=False)
    parser.add_argument("-u", "--NotionDbUrl", help="Notion DB URL", required=False)
    parser.add_argument("-f", "--ForceUpload", help="Force upload to Notion", required=False, action='store_true')
    parser.add_argument("-o", "--OutputFolder", help="Folder where .csv and .xlsx files will be created", required=False)

    args = parser.parse_args()

    if args.Path: path = args.Path
    if args.Extensions:
        supported_extensions = args.Extensions
    else:
        supported_extensions = ['.epub', '.pdf', '.paper']
    if args.NotionToken: notionToken = args.NotionToken
    if args.NotionDbUrl: notionDbUrl = args.NotionDbUrl
    forceUpload = False
    if args.ForceUpload: forceUpload = True
    if args.OutputFolder:
        outputFolder = args.OutputFolder
    else:
        outputFolder = './l2n-output'
    if args.LogFilePath:
        logFilePath = args.LogFilePath
    else:
        logFilePath = outputFolder + '/log.json'


    # ---------------------------------------------------------------------------------------

    print('\nlibrary2notion created by Javier Helguera (github.com/helguera) © 2022 MIT License\n')

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "books"

    ws1.cell(row=1, column=1).value = 'File Name'
    ws1.cell(row=1, column=2).value = 'Title'
    ws1.cell(row=1, column=3).value = 'Priority'
    ws1.cell(row=1, column=4).value = 'Tags'
    ws1.cell(row=1, column=5).value = 'Status'
    ws1.cell(row=1, column=6).value = 'Author'
    ws1.cell(row=1, column=7).value = 'Publisher'
    ws1.cell(row=1, column=8).value = 'Comments'
    ws1.cell(row=1, column=9).value = 'Format'
    ws1.cell(row=1, column=10).value = 'ISBN'


    #####################################################################################################################

    files_path = [os.path.abspath(os.path.join(path, x)) for x in os.listdir(path)]
    files = []
    temp_files = []

    for r, d, f in os.walk(path):
        for file in f:
            if filter(lambda element: file in element, supported_extensions):
                files.append(os.path.join(r, file))
    temp_files = files.copy()
    temp_files = filter(lambda x: '._' not in x, files)
    files = filter(lambda x: '._' not in x, files)

    bookCollection = BookCollection()

    total_files = 0
    for index, my_file in enumerate(temp_files, start=1):
        total_files += 1

    print('\nTOTAL FILES: ' + str(total_files) + '\n')
    printProgressBar(0, total_files, prefix = 'Detecting files:       ', suffix = 'Complete', length = 100)

    for index, my_file in enumerate(files, start=1):

        if os.path.splitext(my_file)[1] in supported_extensions:
            book = Book(os.path.splitext(my_file)[0], [os.path.splitext(my_file)[1].replace('.','').upper()], os.path.splitext(my_file)[1])
            bookCollection.insert(book)

        printProgressBar(index, total_files, prefix = 'Detecting files:       ', suffix = 'Complete', length = 100)

    # JSON PROCESSING
    deleted = []
    if ('logFilePath' in locals() and os.path.exists(logFilePath) and not forceUpload):
        with open(logFilePath, 'r') as openfile:
            # Reading from json file
            json_object = json.load(openfile)
            printProgressBar(0, len(json_object)-1, prefix = 'Analazing log file:    ', suffix = 'Complete', length = 100)
            for index, value in enumerate(json_object):
                bookInCollection = bookCollection.findBook(value['filename'])
                if bookInCollection is not None:
                    if collections.Counter(bookInCollection.formats) == collections.Counter(value['formats']):
                        bookInCollection.action = "do nothing"
                    else:
                        json_object[index]['formats']= bookInCollection.formats
                        bookInCollection.action = "update"
                else:
                    json_object[index]['action']= 'DELETED'
                    deleted.append(json_object[index])
                printProgressBar(index, len(json_object)-1, prefix = 'Analazing log file:    ', suffix = 'Complete', length = 100)

        deleted_json_object = json.dumps(deleted, indent=2)
        with open(os.path.join(outputFolder, 'deleted.json'), "w") as outfile:
            outfile.write(deleted_json_object)

    json_object = json.dumps(bookCollection.createJson(), indent=2)
    if not os.path.exists(outputFolder):
        os.makedirs(outputFolder)
    with open(logFilePath, "w") as outfile:
        outfile.write(json_object)

    # bookCollection.printAll()
    bookCollection.extractMetadataFromBooks(forceUpload)
    bookCollection.copyAllToExcel(ws1)
    if not os.path.exists(outputFolder):
        os.mkdir(outputFolder)
    wb.save(filename=os.path.join(outputFolder, 'books.xlsx'))
    df = pd.read_excel(os.path.join(outputFolder, 'books.xlsx'), 'books', index_col=None)
    df.to_csv (os.path.join(outputFolder, 'books.csv'), index = None, header=True, encoding='utf-8-sig')
    printProgressBar(0, 0, prefix = 'Converting to .csv:    ', suffix = 'Complete', length = 100)

    if not bookCollection.isEmpty() or forceUpload:
        if 'notionToken' in locals() and 'notionDbUrl' in locals():
            print('\n')
            command = 'csv2notion --token ' + notionToken + ' --url ' + notionDbUrl + ' --merge ' + os.path.join(outputFolder, 'books.csv') + ' --merge-only-column "File Name" --merge-only-column "Title" --merge-only-column "Author" --merge-only-column "Publisher" --merge-only-column "Format" --merge-only-column "Tags" --merge-only-column "ISBN" --column-types "text, select, multi_select, select, text, text, text, multi_select, text" --add-missing-columns --verbose'
            os.system(command)
        else:
            print('\n\nCSV file has been created but no Notion data was provided')
    else:
        print('\n\nEverything up to date. Nothing to upload to Notion.')

    print('\nDONE.\n')